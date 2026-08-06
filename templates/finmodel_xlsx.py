#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel-версия финансовой модели (артефакт A7, вторая форма).

Зачем. Руководителю нужна модель, в которой видно, КАК считается результат:
открыл ячейку — увидел формулу — проследил вычисление до исходных параметров.
Поэтому рядом с расчётным скриптом (finmodel.py) пайплайн отдаёт книгу Excel
с ЖИВЫМИ формулами — без единого захардкоженного результата.

Листы:
  Константы  — единственный источник входных значений: все параметры модели
               именованными ячейками (реестр допущений целиком).
  Гипотезы   — срез реестра с меткой ГИПОТЕЗА: что взято на веру, с каким
               основанием и в каких границах варьируется. Значения — ссылки
               на «Константы», не копии.
  Бенчмарки  — срез с метками БЕНЧМАРК и АНАЛОГ: на что опираемся, откуда цифра.
  Модель     — вычисления формулами: годовой эффект → денежный поток по годам →
               дисконтирование → NPV → накопленный поток → окупаемость → сценарии.
  График NPV — диаграмма: столбики потока по годам + линия накопленного
               дисконтированного потока (точка пересечения нуля = окупаемость).

Допущения НЕ дублируются: скрипт импортирует их из finmodel.py, поэтому
Excel-книга и текстовый расчёт всегда согласованы. Рабочий порядок агента:
  1) правишь блок ДОПУЩЕНИЯ в finmodel.py;
  2) python3 finmodel.py            — текстовый отчёт, Монте-Карло, график;
  3) python3 finmodel_xlsx.py       — книга finmodel.xlsx с формулами.

Монте-Карло в Excel-книгу не переносится (это симуляция, а не формула) —
его результаты остаются в отчёте finmodel.py; в листе «Модель» для них
отведён блок «вписать из отчёта», помеченный явно.

Зависимости: openpyxl (pip install openpyxl).
"""

import argparse
import importlib.util
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.workbook.defined_name import DefinedName
except ImportError:  # pragma: no cover
    raise SystemExit("Нужен openpyxl:  pip install openpyxl")


# ── загрузка допущений из finmodel.py (единый источник) ──────────────────────

def load_model(path: Path):
    spec = importlib.util.spec_from_file_location("finmodel", path)
    m = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(m)
    return m


# ── оформление ───────────────────────────────────────────────────────────────

INK = "FF1E2027"
ACCENT = "FFFF5533"
HEAD_FILL = PatternFill("solid", fgColor="FFF2F2F2")
HYP_FILL = PatternFill("solid", fgColor="FFFFF3E0")   # гипотезы — тёплая подсветка
BENCH_FILL = PatternFill("solid", fgColor="FFE8F1FB")  # бенчмарки — холодная
THIN = Side(style="thin", color="FFCCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
H1 = Font(bold=True, size=14, color=INK)
HD = Font(bold=True, size=10, color=INK)
MONEY = "# ##0"


def header(ws, row, cols):
    for c, text in enumerate(cols, start=1):
        cell = ws.cell(row=row, column=c, value=text)
        cell.font = HD
        cell.fill = HEAD_FILL
        cell.border = BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def autowidth(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── сборка книги ─────────────────────────────────────────────────────────────

def build(fm, out_path: Path):
    wb = Workbook()

    имена = list(fm.ДОПУЩЕНИЯ.keys())
    лет = int(fm.ДОПУЩЕНИЯ["горизонт_лет"][0])
    if лет != 3:
        # лист «Модель» разложен по явным строкам лет; при другом горизонте
        # агент обязан пересобрать блок лет руками — лучше упасть громко
        raise SystemExit(f"Книга собирается для горизонта 3 года, в модели {лет}. "
                         "Поправь генератор или горизонт.")

    # ── Константы ────────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Константы"
    ws.cell(row=1, column=1, value=f"Константы модели · {fm.ИНИЦИАТИВА}").font = H1
    ws.cell(row=2, column=1, value="Единственный источник входных значений: модель и листы "
                                   "«Гипотезы»/«Бенчмарки» ссылаются сюда формулами.")
    header(ws, 4, ["параметр", "значение", "ед. изм.", "метка", "основание"])
    строка_конст = {}
    r = 5
    for имя in имена:
        знач, ед, метка, осн = fm.ДОПУЩЕНИЯ[имя]
        ws.cell(row=r, column=1, value=имя).border = BORDER
        c = ws.cell(row=r, column=2, value=знач)
        c.border = BORDER
        c.number_format = MONEY if isinstance(знач, (int,)) and abs(знач) >= 1000 else "General"
        ws.cell(row=r, column=3, value=ед).border = BORDER
        m = ws.cell(row=r, column=4, value=метка)
        m.border = BORDER
        if метка == "ГИПОТЕЗА":
            m.fill = HYP_FILL
        elif метка in ("БЕНЧМАРК", "АНАЛОГ"):
            m.fill = BENCH_FILL
        ws.cell(row=r, column=5, value=осн).border = BORDER
        # именованная ячейка: в формулах модели читается как «п_имя»
        wb.defined_names.add(DefinedName(f"п_{имя}", attr_text=f"Константы!$B${r}"))
        строка_конст[имя] = r
        r += 1
    autowidth(ws, [24, 16, 11, 12, 58])

    # ── Гипотезы / Бенчмарки: срезы реестра ссылками ────────────────────────
    def срез(title, метки, fill, пояснение):
        w = wb.create_sheet(title)
        w.cell(row=1, column=1, value=title).font = H1
        w.cell(row=2, column=1, value=пояснение)
        header(w, 4, ["параметр", "значение (ссылка)", "ед. изм.", "метка", "основание"])
        rr = 5
        for имя in имена:
            _, ед, метка, осн = fm.ДОПУЩЕНИЯ[имя]
            if метка not in метки:
                continue
            k = строка_конст[имя]
            w.cell(row=rr, column=1, value=имя).border = BORDER
            f = w.cell(row=rr, column=2, value=f"=Константы!B{k}")
            f.border = BORDER
            f.fill = fill
            w.cell(row=rr, column=3, value=f"=Константы!C{k}").border = BORDER
            w.cell(row=rr, column=4, value=f"=Константы!D{k}").border = BORDER
            w.cell(row=rr, column=5, value=f"=Константы!E{k}").border = BORDER
            rr += 1
        autowidth(w, [24, 18, 11, 12, 58])
        return w, rr

    w_hyp, rr = срез(
        "Гипотезы", {"ГИПОТЕЗА"}, HYP_FILL,
        "Что взято на веру (погрешность до 50 % — принятый в методике порог). Проверяется пилотом; менять — на листе «Константы».")
    # границы варьирования из РАЗБРОС — для гипотез и стресс-параметров
    w_hyp.cell(row=rr + 1, column=1, value="Границы варьирования (консервативно · базово · оптимистично):").font = HD
    hr = rr + 2
    header(w_hyp, hr, ["параметр", "консервативно", "базово", "оптимистично", ""])
    строка_разброс = {}
    for имя, (lo, mid, hi) in fm.РАЗБРОС.items():
        hr += 1
        w_hyp.cell(row=hr, column=1, value=имя).border = BORDER
        for c, v in ((2, lo), (3, mid), (4, hi)):
            cell = w_hyp.cell(row=hr, column=c, value=v)
            cell.border = BORDER
            cell.number_format = MONEY
        строка_разброс[имя] = hr

    срез("Бенчмарки", {"БЕНЧМАРК", "АНАЛОГ"}, BENCH_FILL,
         "На что опираемся: БЕНЧМАРК — подтверждено измерением, АНАЛОГ — перенос из смежного случая.")

    # ── Модель ───────────────────────────────────────────────────────────────
    w = wb.create_sheet("Модель")
    w.cell(row=1, column=1, value="Модель вычисления").font = H1
    w.cell(row=2, column=1, value="Ни одного захардкоженного числа: каждая ячейка — формула от «Констант». "
                                  "Меняешь константу — модель пересчитывается.")

    w.cell(row=4, column=1, value="Годовой эффект на полном режиме, ₽").font = HD
    эф = w.cell(row=4, column=2, value="=п_заявок_в_год*(п_часов_было-п_часов_стало)*п_стоимость_часа")
    эф.number_format = MONEY
    w.cell(row=4, column=3, value="заявки × сэкономленные часы × стоимость часа")

    w.cell(row=5, column=1, value="Сопровождение в год, ₽").font = HD
    сп = w.cell(row=5, column=2, value="=п_внедрение*п_доля_сопровождения")
    сп.number_format = MONEY

    header(w, 7, ["год", "денежный поток, ₽", "дисконт-множитель", "дисконтир. поток, ₽",
                  "накопленный дисконтир., ₽", "комментарий"])
    потоки = [
        ("0", "=-п_внедрение", "внедрение"),
        ("1", "=$B$4*п_выход_на_режим-$B$5", "эффект × выход на режим − сопровождение"),
        ("2", "=$B$4-$B$5", "полный эффект − сопровождение"),
        ("3", "=$B$4-$B$5", "полный эффект − сопровождение"),
    ]
    for i, (год, f, комм) in enumerate(потоки):
        rr = 8 + i
        w.cell(row=rr, column=1, value=int(год)).border = BORDER
        cf = w.cell(row=rr, column=2, value=f)
        cf.border = BORDER
        cf.number_format = MONEY
        dm = w.cell(row=rr, column=3, value=f"=1/(1+п_ставка)^A{rr}")
        dm.border = BORDER
        dm.number_format = "0.000"
        dcf = w.cell(row=rr, column=4, value=f"=B{rr}*C{rr}")
        dcf.border = BORDER
        dcf.number_format = MONEY
        cum = w.cell(row=rr, column=5,
                     value=f"=D{rr}" if i == 0 else f"=E{rr - 1}+D{rr}")
        cum.border = BORDER
        cum.number_format = MONEY
        w.cell(row=rr, column=6, value=комм).border = BORDER

    w.cell(row=13, column=1, value="NPV, ₽").font = Font(bold=True, size=12, color=ACCENT)
    npv = w.cell(row=13, column=2, value="=SUM(D8:D11)")
    npv.font = Font(bold=True, size=12)
    npv.number_format = MONEY
    w.cell(row=13, column=3, value="сумма дисконтированных потоков (= накопленный на год 3)")

    w.cell(row=14, column=1, value="Окупаемость (простая), мес").font = HD
    w.cell(row=14, column=2,
           value="=IF(SUM(B8:B11)<0,\"не окупается\","
                 "IF(SUM($B$8:$B$9)>=0,ROUND((-B8/B9)*12,0),"
                 "IF(SUM($B$8:$B$10)>=0,ROUND((1+(-SUM($B$8:$B$9)/B10))*12,0),"
                 "ROUND((2+(-SUM($B$8:$B$10)/B11))*12,0))))")
    w.cell(row=14, column=3, value="по недисконтированному потоку, как в finmodel.py; "
                                   "внутри года — линейная интерполяция")

    # сценарии: формулы от границ на листе «Гипотезы»
    w.cell(row=16, column=1, value="Сценарии").font = HD
    header(w, 17, ["сценарий", "заявок/год", "часов стало", "внедрение, ₽", "NPV, ₽", ""])
    if set(fm.РАЗБРОС.keys()) != {"заявок_в_год", "часов_стало", "внедрение"}:
        raise SystemExit("Сценарный блок книги рассчитан на разброс по "
                         "(заявок_в_год, часов_стало, внедрение) — обнови генератор под новый РАЗБРОС.")
    сц_кол = {"заявок_в_год": None, "часов_стало": None, "внедрение": None}
    for i, (сц, col) in enumerate((("консервативный", 2), ("базовый", 3), ("оптимистичный", 4))):
        rr = 18 + i
        w.cell(row=rr, column=1, value=сц).border = BORDER
        refs = {}
        for c, имя in ((2, "заявок_в_год"), (3, "часов_стало"), (4, "внедрение")):
            src = f"Гипотезы!{get_column_letter(col)}{строка_разброс[имя]}"
            cell = w.cell(row=rr, column=c, value=f"={src}")
            cell.border = BORDER
            cell.number_format = MONEY
            refs[имя] = f"{get_column_letter(c)}{rr}"
        эффект = (f"({refs['заявок_в_год']}*(п_часов_было-{refs['часов_стало']})*п_стоимость_часа)")
        сопр = f"({refs['внедрение']}*п_доля_сопровождения)"
        f_npv = (f"=-{refs['внедрение']}"
                 f"+({эффект}*п_выход_на_режим-{сопр})/(1+п_ставка)"
                 f"+({эффект}-{сопр})/(1+п_ставка)^2"
                 f"+({эффект}-{сопр})/(1+п_ставка)^3")
        cell = w.cell(row=rr, column=5, value=f_npv)
        cell.border = BORDER
        cell.number_format = MONEY
        cell.font = HD

    w.cell(row=22, column=1, value="Монте-Карло (из отчёта finmodel.py — вписать значения)").font = HD
    header(w, 23, ["P10, ₽", "P50, ₽", "P90, ₽", "P(NPV>0)", "", ""])
    for c in range(1, 5):
        w.cell(row=24, column=c, value="—").border = BORDER
    w.cell(row=25, column=1, value="Симуляция — не формула: Excel-книга держит детерминированную модель, "
                                   "разброс считает finmodel.py (10 000 итераций — принятый в методике порог).")
    autowidth(w, [26, 20, 18, 20, 24, 46])

    # ── График NPV ───────────────────────────────────────────────────────────
    g = wb.create_sheet("График NPV")
    g.cell(row=1, column=1, value="Когда инициатива окупается").font = H1
    g.cell(row=2, column=1, value="Столбики — денежный поток по годам; линия — накопленный "
                                  "дисконтированный поток. Пересечение нуля = срок окупаемости.")

    bars = BarChart()
    bars.type = "col"
    bars.title = "Денежный поток и накопленный дисконтированный поток"
    bars.height = 12
    bars.width = 26
    data = Reference(w, min_col=2, min_row=7, max_row=11)      # поток по годам
    cats = Reference(w, min_col=1, min_row=8, max_row=11)      # годы
    bars.add_data(data, titles_from_data=True)
    bars.set_categories(cats)
    bars.y_axis.title = "₽"
    bars.x_axis.title = "год"

    line = LineChart()
    ldata = Reference(w, min_col=5, min_row=7, max_row=11)     # накопленный дисконтированный
    line.add_data(ldata, titles_from_data=True)
    line.y_axis.axId = 200
    bars += line
    g.add_chart(bars, "A4")

    wb.save(out_path)
    return out_path


def main():
    ap = argparse.ArgumentParser(description="Сборка Excel-книги финансовой модели из finmodel.py")
    ap.add_argument("--model", default=str(Path(__file__).with_name("finmodel.py")),
                    help="путь к finmodel.py с блоком ДОПУЩЕНИЯ")
    ap.add_argument("-o", "--out", default="finmodel.xlsx", help="куда сохранить книгу")
    args = ap.parse_args()
    fm = load_model(Path(args.model))
    out = build(fm, Path(args.out))
    print(f"OK: {out} — листы: Константы · Гипотезы · Бенчмарки · Модель · График NPV")


if __name__ == "__main__":
    main()
